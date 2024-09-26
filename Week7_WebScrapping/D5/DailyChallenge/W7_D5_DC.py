import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# 1. Create an Excel workbook with a sheet named "Grades"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Grades"

# 2. Data structure with student grades
data = {
    "Joe": {
        "math": 65,
        "science": 78,
        "english": 98,
        "gym": 89
    },
    "Bill": {
        "math": 55,
        "science": 72,
        "english": 87,
        "gym": 95
    },
    "Tim": {
        "math": 100,
        "science": 45,
        "english": 75,
        "gym": 92
    },
    "Sally": {
        "math": 30,
        "science": 25,
        "english": 45,
        "gym": 100
    },
    "Jane": {
        "math": 100,
        "science": 100,
        "english": 100,
        "gym": 60
    }
}

subjects = ["Math", "Science", "English", "Gym"]
headers = ["Name"] + subjects

# 5. Apply bold and colored formatting to the header row
bold_font = Font(bold=True)
fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')

# Write headers with formatting
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = bold_font
    cell.fill = fill

# 3. Add a row for each student's grades
row_num = 2
for student, grades in data.items():
    ws.cell(row=row_num, column=1).value = student  # Student Name
    for col_num, subject in enumerate(subjects, 2):  # Start from column 2
        ws.cell(row=row_num, column=col_num).value = grades[subject.lower()]
    row_num += 1

# 4. Add a formula to calculate the average grade for each subject
ws.cell(row=row_num, column=1).value = "Average"
for col_num in range(2, len(subjects) + 2):
    col_letter = get_column_letter(col_num)
    start_cell = f"{col_letter}2"
    end_cell = f"{col_letter}{row_num - 1}"
    avg_formula = f"=AVERAGE({start_cell}:{end_cell})"
    ws.cell(row=row_num, column=col_num).value = avg_formula

# Save the workbook
wb.save('grades.xlsx')
