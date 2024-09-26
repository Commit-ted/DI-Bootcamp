import sys
from openpyxl import load_workbook

# Add custom path to system path if needed
# sys.path.append(r"C:\wiseowl\Python\Courseware\Lib\site-packages")

# Work with Excel - updated file path for your Plants.xlsx file
plant_book = load_workbook(filename=r"/Users/manuel/Desktop/DI-Bootcamp/Week7_WebScrapping/D5/ExerciseXP/EX_2/Plants.xlsx")

# Get reference to the first sheet (assuming it's called 'Sheet1')
plant_sheet = plant_book["Sheet1"]

# Start at the second row, first column (A2), assuming header row exists
this_plant = plant_sheet.cell(2, 1)

# Initialize a safeguard counter to avoid infinite loops
int_check = 0

# Loop down cells in the first column until an empty cell is found
while True:
    # Avoid infinite loops if bug
    int_check += 1
    if int_check > 100:  # Adjust if you expect more than 100 rows
        print("Integer check triggered. Exiting the loop.")
        break

    # Go to the next plant (move down one row)
    this_plant = this_plant.offset(1, 0)

    # If there is no plant (cell is empty), stop
    if this_plant.value is None:
        print("\nThe above plants are not in stock")
        break

    # Check the 'In Stock' column (which is 7 columns to the right of the plant name, i.e., column H)
    stock_cell = this_plant.offset(0, 7)  # 'H' column corresponds to offset(0, 7)

    # If the plant is not in stock ('No'), print its name
    if stock_cell.value == "No":
        print(this_plant.value)

# Close the workbook
plant_book.close()

