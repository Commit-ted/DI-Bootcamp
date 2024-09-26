import pandas as pd
from openpyxl import load_workbook

# File path to your Excel file
file_path = "/Users/manuel/Desktop/DI-Bootcamp/Week7_WebScrapping/D5/ExerciseXP/EX_3_Data.xlsx"

# Load the Excel file into a DataFrame
df = pd.read_excel(file_path)

# Perform data manipulation: filter rows where 'Sales' is greater than 1000
filtered_df = df[df['Sales'] > 1000]

# Load the existing workbook using openpyxl
workbook = load_workbook(filename=file_path)

# Access the active sheet (or provide the sheet name if necessary)
sheet = workbook.active

# Remove any existing data from the worksheet (optional, depending on your needs)
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
    for cell in row:
        cell.value = None

# Write the filtered DataFrame back to the same Excel file (starting from the second row)
for index, row in filtered_df.iterrows():
    for col_index, value in enumerate(row):
        sheet.cell(row=index + 2, column=col_index + 1, value=value)

# Save the updated workbook
workbook.save(filename=file_path)

print(f"Filtered data has been written back to {file_path}")
