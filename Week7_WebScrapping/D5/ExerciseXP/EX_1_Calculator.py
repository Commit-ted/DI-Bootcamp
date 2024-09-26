import openpyxl

# Create a new Excel workbook
wb = openpyxl.Workbook()
ws = wb.active

# Write labels in the first column
ws['A1'] = 'First number ==>'
ws['A2'] = 'Second number ==>'
ws['A3'] = 'Result ==>'

# Set up the formula for multiplication in the Result cell
ws['B3'] = '=B1*B2'

# Save the Excel file in the same directory
wb.save('calculator_EX_1.xlsx')
